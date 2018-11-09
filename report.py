# -*- coding: UTF-8 -*-
import getopt
import sys
import pymysql
import openpyxl
import time
import os
import datetime


class Main(object):
    short_args = ''
    long_args = ['product=', 'date=']
    opts, args = getopt.getopt(sys.argv[1:], short_args, long_args)

    data2 = {
            '后台管理Web': 1,
            '运营客服管理Web': 2,
            '卖家中介版Android': 3,
            '卖家中介版OS': 4,
            '地推版Android': 5,
            '买家版Web': 6,
            '买家版iOS': 7,
            '买家版Android': 8,
            '卖家中介v1.0.6': 9,
            '用户版Android': 10,
            '用户版iOS': 11,
            '用户版Web': 12
            }

    def __init__(self):
        for opt, value in self.opts:
            self.date = time.strftime('%Y-%m-%d', time.localtime(time.time()))
            self.product = '用户版Web'
            if opt == '--product':
                self.product = value
            elif opt == '--date':
                self.date = str(datetime.datetime.date(datetime.datetime.strptime(value, '%Y-%m-%d')))
            else:
                print('参数 --%s %s 无效') % (opt, value)

    def get_product_id(self):
        for k, v in self.data2.items():
            # print(k, v)
            if self.product == k:
                return v
        return False


class MysqlDb1(object):

    # 连接数据库
    def conn(self, host, port, dbname, user, password):
        try:
            conn = pymysql.connect(host, user, password, dbname, port, charset='utf8')
            return conn
        except Exception as e:
            print("connect failed", e)

    # 查询
    def sqlsearch(self, sql, db):
        try:
            cur = db.cursor()
            x = cur.execute(sql)                   # 使用cursor进行各种操作
            results = cur.fetchmany(x)
            cur.close()
            return results
        except Exception as e:
            print("search failed", e)

    # 直接增删改
    def sqlDML(self, sql, db):
        # include: insert,update,delete
        try:
            cr = db.cursor()
            cr.execute(sql)
            cr.close()
            db.commit()
        except Exception as e:
            print("sqlDML failed",e)

    # 有参数增删改
    def sqlDML2(self, sql, params, db):
        # execute dml with parameters
        try:
            cr = db.cursor()
            cr.executemany(sql, params)
            cr.close()
            db.commit()
        except Exception as e:
            print("sqlDML2 failed", e)


def create_xlsx(filename):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'test_report'
    wb.save(filename=filename)
    print("新建Excel："+filename+"成功")


def savetoexcel(create_today_list, closed_today_list, rest_bug_list, create_today_all, bug_detail, sheetname, wbname):
    wb = openpyxl.load_workbook(filename=wbname)

    sheet = wb.active
    sheet.title = sheetname

    sheet['B7'] = bug_detail

    for col in range(3, 7):
        flg = 0
        if create_today_list is not None:
            for i in create_today_list:
                if col == i[0] + 2:
                    flg = 1
                    _ = sheet.cell(row=10, column=col, value=i[1])
            if flg == 0:
                _ = sheet.cell(row=10, column=col, value=0)
        else:
            _ = sheet.cell(row=10, column=col, value=0)

    for col in range(3, 7):
        flg = 0
        if closed_today_list is not None:
            for i in closed_today_list:
                if col == i[0] + 2:
                    flg = 1
                    _ = sheet.cell(row=11, column=col, value=i[1])
            if flg == 0:
                _ = sheet.cell(row=11, column=col, value=0)
        else:
            _ = sheet.cell(row=11, column=col, value=0)

    for col in range(3, 7):
        flg = 0
        if rest_bug_list is not None:
            for i in rest_bug_list:
                if col == i[0] + 2:
                    flg = 1
                    _ = sheet.cell(row=12, column=col, value=i[1])
            if flg == 0:
                _ = sheet.cell(row=12, column=col, value=0)
        else:
            _ = sheet.cell(row=12, column=col, value=0)

    sheet['C13'] = create_today_all

    # border = Border(left=Side(style='medium', color='FF000000'), right=Side(style='medium', color='FF000000'),
    #                 top=Side(style='medium', color='FF000000'), bottom=Side(style='medium', color='FF000000'),
    #                 diagonal=Side(style='medium', color='FF000000'), diagonal_direction=0,
    #                 outline=Side(style='medium', color='FF000000'), vertical=Side(style='medium', color='FF000000'),
    #                 horizontal=Side(style='medium', color='FF000000'))

    wb.save(filename=wbname)


def replace_xls(sheetname, filename, filename2):

    wb = openpyxl.load_workbook(filename)
    wb2 = openpyxl.load_workbook(filename2)

    ws = wb[sheetname]
    ws2 = wb2[sheetname]

    #两个for循环遍历整个excel的单元格内容
    for i,row in enumerate(ws.iter_rows()):
        for j,cell in enumerate(row):
            ws2.cell(row=i+1, column=j+1, value=cell.value)

    wb2.save(filename2)


if __name__ == '__main__':
    user = 'admin'
    password = '123456a'
    host = "192.168.88.205"
    port = 3306
    db_name = 'zentao'
    datatime = time.strftime('%Y-%m-%d', time.localtime(time.time()))

    data1 = Main()
    pro_id = data1.get_product_id()
    print(data1.product, data1.date, data1.get_product_id())

    time1 = "'" + data1.date + "'"
    sql_create_today = "select severity,count(*) from zt_bug WHERE DATE_FORMAT(openedDate,'%%y-%%m-%%d')  = DATE_FORMAT(%s,'%%y-%%m-%%d')\
                        and product = %s GROUP BY severity;" % (time1, str(pro_id))
    # print(sql_create_today)
    sql_closed_today = "select severity,count(*) from zt_bug WHERE DATE_FORMAT(closedDate,'%%y-%%m-%%d')  = DATE_FORMAT(%s,'%%y-%%m-%%d') \
                        and product = '%s' GROUP BY severity;" % (time1, str(pro_id))
    sql_rest_bug = "select severity,count(*) from zt_bug WHERE product = '%s' and status != 'closed' \
                    GROUP BY severity;" % str(pro_id)
    sql_create_today_all = "select count(*) from zt_bug WHERE DATE_FORMAT(openedDate,'%%y-%%m-%%d')  = DATE_FORMAT(%s,'%%y-%%m-%%d') \
                            and product = '%s';" % (time1, str(pro_id))
    sql_bug_list = "select id,severity,pri,resolution,resolvedBuild,title,assignedTo from zt_bug \
                    WHERE product = '%s' and status != 'closed' order BY severity LIMIT 10;" % str(pro_id)
    sqldb = MysqlDb()
    conn = sqldb.conn(host, port, db_name, user, password)
    create_today_list = sqldb.sqlsearch(sql_create_today, conn)
    closed_today_list = sqldb.sqlsearch(sql_closed_today, conn)
    rest_bug_list = sqldb.sqlsearch(sql_rest_bug, conn)
    create_today_all = sqldb.sqlsearch(sql_create_today_all, conn)
    # print(create_today_all)
    if create_today_all is not None:
        create_today_all = create_today_all[0][0]
    else:
        create_today_all = 0
    bug_list = list(sqldb.sqlsearch(sql_bug_list, conn))
    # print(create_today_list)
    # print(closed_today_list)
    # print(rest_bug_list)
    # print(create_today_all)
    # print(bug_list)
    bug_detail = ''
    if bug_list is not None:
        for bug in bug_list:
            bug = list(bug)
            if bug[3] == '':
                bug[3] = '激活'
            else:
                bug[3] = '已解决'
            if bug[4] == '':
                bug[4] = '未确认'
            else:
                bug[4] = '已确认'
            bug[5] = bug[5].replace('&quot;', '')
            for detail in bug:
                if detail != bug[-1]:
                    bug_detail = bug_detail + '[' + str(detail) + ']-'
                else:
                    bug_detail = bug_detail + '[' + str(detail) + ']'
            bug_detail = bug_detail + '\n'

    print(bug_detail)

    fileName = '模板.xlsx'
    suffixPosition = fileName.rfind(".")
    newFileName = fileName[:suffixPosition] + str(data1.date) + fileName[suffixPosition:]
    # print(newFileName)
    # print(os.path.exists(newFileName))
    if os.path.exists(newFileName) is False:
        create_xlsx(newFileName)
    # 拷贝
    newFile = open(newFileName, "wb")
    oldFile = open(fileName, "rb")
    contents = oldFile.readlines()
    newFile.writelines(contents)

    newFile.close()
    oldFile.close()

    # replace_xls('test_report', fileName, newFileName)

    savetoexcel(create_today_list, closed_today_list, rest_bug_list, create_today_all, bug_detail,
                'test_report', newFileName)
