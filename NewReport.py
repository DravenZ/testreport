# encoding: utf-8

"""
__author__ = "Zhang Pengfei"
__date__ = 2018/11/7
"""


import openpyxl
from QueryData import QueryData
import datetime
from openpyxl.styles import PatternFill, Alignment, Border, Side, Font
from openpyxl.chart import (
    Reference,
    Series,
    BarChart3D,
)

font = Font(size=12)
BUG_GRADE = ["阻塞缺陷", "一般缺陷", "显示缺陷", "建议缺陷"]
fill = PatternFill("solid", fgColor="00B050")
fill1 = PatternFill("solid", fgColor="FF0000")
fill2 = PatternFill("solid", fgColor="FFC000")
fill3 = PatternFill("solid", fgColor="FFFF00")
fill4 = PatternFill("solid", fgColor="A6A6A6")
column_width = 15
alignment = Alignment(horizontal='center', vertical='center')
now_date = datetime.datetime.now().date()
border = Border(left=Side(style='medium', color='A6A6A6'), right=Side(style='medium', color='A6A6A6'),
                top=Side(style='medium', color='A6A6A6'), bottom=Side(style='medium', color='A6A6A6'),
                diagonal=Side(style='medium', color='A6A6A6'), diagonal_direction=0,
                outline=Side(style='medium', color='A6A6A6'), vertical=Side(style='medium', color='A6A6A6'),
                horizontal=Side(style='medium', color='A6A6A6'))

world_farm = {
    2: '运营客服',
    10: '世界农场',
    14: '苗叔运营后台',
    15: '苗叔用户端'
}


def create_xlsx(file_name):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'test_report'
    wb.save(filename=file_name)
    print("新建Excel："+file_name+"成功")


def write_xlsx(*args, **kwargs):
    for arg in args:
        print(arg)
    # print(args[7])
    wb = openpyxl.load_workbook(filename=args[0])
    sheet = wb.active
    # 设置行列宽
    sheet.column_dimensions['C'].width = column_width
    sheet.column_dimensions['D'].width = column_width
    sheet.column_dimensions['E'].width = column_width
    sheet.column_dimensions['F'].width = column_width
    sheet.column_dimensions['G'].width = column_width
    sheet.column_dimensions['L'].width = 150
    sheet.column_dimensions['k'].width = column_width
    sheet.column_dimensions['M'].width = column_width
    sheet.column_dimensions['N'].width = column_width
    sheet.row_dimensions[6].height = 20
    sheet.row_dimensions[16].height = 20

    # 第一和表格的填充
    min_row = 6
    max_row = min_row + len(list(args[2])) + 1
    min_col = 3
    max_col = min_col + 1 + len(args[1])
    cs = sheet.cell(row=min_row, column=min_col)
    cs.fill = fill
    cs.border = border
    for col in range(min_col+1, max_col):
        cs = sheet.cell(row=min_row, column=col)
        cs.value = args[1][col-min_col - 1]
        cs.fill = fill
        cs.alignment = alignment
        cs.font = font

    for row in range(min_row + 1, max_row):
        cs = sheet.cell(row=row, column=min_col)
        cs.value = list(args[2])[row - min_row - 1]
        cs.fill = fill
        cs.alignment = alignment
        cs.font = font

    for row in range(min_row + 1, max_row):
        for v in args[3][row - min_row - 1]:
            cs = sheet.cell(row=row, column=list(v)[0] + min_col)
            cs.value = list(v)[1]
            cs.alignment = alignment
            cs.font = font

    for row in range(min_row, max_row):
        sheet.row_dimensions[row].height = 25
        for col in range(min_col, max_col):
                cs = sheet.cell(row=row, column=col)
                cs.border = border
                cs.font = font

    # # 第一个图表
    # data = Reference(sheet, min_col=3, min_row=6, max_col=7, max_row=10)
    # titles = Reference(sheet, min_col=4, min_row=6, max_col=7)
    # # tuli = Reference(sheet, min_col=3, min_row=7, max_row=10)
    # chart = BarChart3D()
    # chart.title = "平台分布"
    # chart.add_data(data=data, titles_from_data=True)
    # chart.set_categories(titles)
    # sheet.add_chart(chart, "H6")

    max_row = sheet.max_row
    min_row = max_row + 5
    max_row = min_row + len(args[4]) + 1
    # 第二个表格的填充
    cs = sheet.cell(row=min_row, column=min_col)
    cs.fill = fill
    for col in range(min_col + 1, max_col):
        cs = sheet.cell(row=min_row, column=col)
        cs.value = args[1][col - min_col - 1]
        cs.fill = fill
        cs.alignment = alignment
        cs.font = font

    for row in range(len(args[4])):
        cs = sheet.cell(row=row + min_row + 1, column=min_col)
        cs.value = list(args[4])[row][0]
        cs.fill = fill
        cs.alignment = alignment
        cs.font = font

    for col in range(min_col + 1, max_col):
        for i in args[5][col - min_col - 1]:
            for row in range(len(args[4])):
                if list(i)[0] == list(args[4])[row][0]:
                    cs = sheet.cell(row=row + min_row + 1, column=col)
                    cs.value = list(i)[1]
                    cs.alignment = alignment
                    cs.font = font

    for row in range(min_row, min_row + 1 + len(args[4])):
        sheet.row_dimensions[row].height = 25
        for col in range(min_col, max_col):
                cs = sheet.cell(row=row, column=col)
                cs.border = border
                cs.font = font
    # # 第一个图表
    # data = Reference(sheet, min_row=6, min_col=4, max_row=10, max_col=8)
    # titles = Reference(sheet, min_col=3, min_row=7, max_row=10)
    # # tuli = Reference(sheet, min_col=3, min_row=7, max_row=10)
    # chart = BarChart3D()
    # chart.title = "严重程度分布"
    # chart.add_data(data=data, titles_from_data=True)
    # chart.set_categories(titles)
    # sheet.add_chart(chart, "H17")

    max_row = sheet.max_row
    min_row = max_row + 5
    max_row = min_row + 1 + len(list(args[2]))
    # 第三个表格的填充
    cs = sheet.cell(row=min_row, column=min_col)
    cs.fill = fill
    for col in range(min_col + 1, max_col):
        cs = sheet.cell(row=min_row, column=col)
        cs.value = str(now_date + datetime.timedelta(days=- (max_col-1-col)))
        cs.fill = fill
        cs.alignment = alignment
        cs.font = font

    for row in range(min_row + 1, max_row):
        cs = sheet.cell(row=row, column=min_col)
        cs.value = list(args[2])[row - (min_row + 1)]
        cs.fill = fill
        cs.alignment = alignment
        cs.font = font

    for col in range(min_col + 1, max_col):
        for row in range(min_row + 1, max_row):
            cs = sheet.cell(row=row, column=col)
            cs.value = list(list(args[6][col - min_col - 1][row-(min_row + 1)])[0])[0]
            cs.alignment = alignment
            cs.font = font

    for row in range(min_row, max_row):
        sheet.row_dimensions[row].height = 25
        for col in range(min_col, max_col):
                cs = sheet.cell(row=row, column=col)
                cs.border = border
                cs.font = font

    max_row = sheet.max_row

    # 第四个表格的填充
    table_title = ['项目', "缺陷标题", '验证程度', '指派给']
    min_col = 11
    max_col = min_col + len(table_title)
    min_row = 6

    cs = sheet.cell(row=min_row, column=min_col)
    cs.fill = fill
    for col in range(min_col, max_col):
        cs = sheet.cell(row=min_row, column=col)
        cs.value = table_title[col-max_col]
        cs.fill = fill
        cs.alignment = alignment
        cs.border = border
        cs.font = font

    flag = 0
    for row in range(len(args[7])):
        sheet.row_dimensions[row+min_row].height = 25
        for col in range(min_col, max_col):
            if world_farm.get(args[7][row][0]) is not None and col == 11:
                cs = sheet.cell(row=row + min_row + 1, column=col)
                cs.value = world_farm.get(list(args[7][row])[0])
            else:
                cs = sheet.cell(row=row + min_row + 1, column=col)
                cs.value = args[7][row][col-min_col]
            if col == max_col - 2:
                if args[7][row][2] == 1:
                    cs.value = BUG_GRADE[0]
                    cs.fill = fill1
                elif args[7][row][2] == 2:
                    cs.value = BUG_GRADE[1]
                    cs.fill = fill2
                elif args[7][row][2] == 3:
                    cs.value = BUG_GRADE[2]
                    cs.fill = fill3
                elif args[7][row][2] == 4:
                    cs.value = BUG_GRADE[3]
                    cs.fill = fill4
            if col == max_col - 1:
                if args[7][row][2] == 1:
                    cs.fill = fill1
                elif args[7][row][2] == 2:
                    cs.fill = fill2
                elif args[7][row][2] == 3:
                    cs.fill = fill3
                elif args[7][row][2] == 4:
                    cs.fill = fill4
            if col != min_col + 1:
                cs.alignment = alignment
            if col == min_col:
                cs.fill = fill
            if col == min_col + 1:
                link = 'http://bug.sjnc.com/index.php?m=bug&f=view&bugID=' + str(args[7][row][4])
                cs.hyperlink = (link)
            cs.border = border
            cs.font = font
    for row in range(len(args[7])):
        if args[7][row][0] != args[7][row-1][0] and row != 0:
            str1 = 'K' + str(flag+min_row+1) + ":K" + str(row-1 + min_row+1)
            sheet.merge_cells(str1)
            flag = row
        if row == len(args[7])-1:
            str1 = 'K' + str(flag + min_row+1) + ":K" + str(row + min_row+1)
            sheet.merge_cells(str1)
            sheet.row_dimensions[row + min_row+1].height = 25
    wb.save(filename=args[0])


if __name__ == '__main__':
    product = 2
    Q = QueryData(product)
    filename = str(product) + str(Q.now_date) + "测试报告.xlsx"
    create_xlsx(filename)
    write_xlsx(filename, BUG_GRADE, Q.duan.keys(), Q.query_create_all(), Q.query_personnel_list(),
               Q.query_bug_list_by_person(), Q.query_bug_by_date(), Q.query_bug_list())




